from certificate_module import *
from openpyxl import load_workbook

load_wb = load_workbook('certificate.xlsx') # 엑셀 파일 읽어오기
load_ws = load_wb.active # 읽어온 엑셀 파일에서 활성화된 시트 불러오기

name_list = []
birth_date_list = []
number_list = []

for i in range(1, load_ws.max_row + 1):
    name_list.append(load_ws.cell(i, 1).value)
    birth_date_list.append(load_ws.cell(i, 2).value)
    number_list.append(load_ws.cell(i, 3).value)

print(name_list)
print(birth_date_list)
print(number_list)


contents = {
    "certificate_title" : "수 료 증",
    "courses" : "Python 1, Python 2",
    "education_date" : "2025.03.09 ~ 2025.05.03",
    "award_statement_1" : "위 사람은",
    "award_statement_2" : "교육과정을 탁월하게 이수하였으므로 이 증서를 수여 합니다.",
    "academy_name" : "코리아 IT 아카데미"
}

for i in range(len(name_list)):
    output_file_name = f"certificate_{name_list[i]}_{i+1}"
    form_file_name = "certificate_form"
    # make_certificate(name_list[i], birth_date_list[i], number_list[i], form_file_name, output_file_name, certificate_title="수 료 증", courses="Python 1, Python 2", education_date="2025.03.09 ~ 2025.05.03")
    make_certificate(name_list[i], birth_date_list[i], number_list[i], form_file_name, output_file_name, **contents)

