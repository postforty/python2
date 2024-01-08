# 엑셀 파일의 주소로 위경도 얻어 엑셀 파일 생성

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl import Workbook
import re

path = '고등교육기관 하반기 주소록(2022).xlsx'
df_excel = pd.read_excel(path, engine='openpyxl')
df_excel.columns = df_excel.loc[4].tolist()
df_excel = df_excel.drop(index=list(range(0, 5)))

def request_geo(road):
    apiurl = "http://api.vworld.kr/req/address?"
    params = {
        "service": "address",
        "request": "getcoord",
        "crs": "epsg:4326",
        "address": road,
        "format": "json",
        "type": "road",
        "key": "CABC17B4-3089-3C88-8FD5-4BEB23099EDE"
    }

    response = requests.get(apiurl, params=params)
    json_data = response.json()

    # print(json_data)

    # return(json_data)
    if json_data['response']['status'] == 'OK':
        x = json_data['response']['result']['point']['x']
        y = json_data['response']['result']['point']['y']
        return x, y
    else:
        x = 0
        y = 0
        return x, y

# "학교주소좌표.xlsx" 파일을 읽어오는 코드, 해당 파일이 존재하지 않으면 에러가 발생하기 때문에 예외 처리함
wb = load_workbook('학교주소좌표.xlsx', data_only=True)
sheet = wb.active

try:
    # '학교주소좌표.xlsx'파일이 존재하면 읽어서 워크시트 활성화
    wb = load_workbook('학교주소좌표.xlsx', data_only=True)
    sheet = wb.active
except:
    # Workbook 초기화
    wb = Workbook()
    sheet = wb.active

# print(df_excel)
name_list = df_excel['학교명'].to_list()
# print(name_list)
address_list = df_excel['주소'].to_list()

print(address_list)
for i, v in enumerate(address_list):
    # print(i)
    # print(v)
    # print(request_geo(v))
    new_address = re.sub('\([^)]*|\)', '', v) # "("로 시작해서 ")"가 포함되지 않는 0개 이상의 모든 문자 또는 )를 ''으로 만듦
    # print(new_address)
    x, y = request_geo(new_address)
    sheet.append([name_list[i], new_address, x, y]) # 시트 내용 추가

# 추가된 시트 내용 보기 
from pandas import DataFrame
df = DataFrame(sheet.values)
print(df)

wb.save("학교주소좌표.xlsx")