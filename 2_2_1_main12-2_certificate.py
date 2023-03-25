from openpyxl import load_workbook

load_wb = load_workbook('certificate.xlsx') # 엑셀 파일 읽어오기
load_ws = load_wb.active # 읽어온 엑셀 파일에서 활성화된 시트 불러오기

# print(dir(load_ws))
# print(load_ws.max_column)
# print(load_ws.max_row)
# print(load_ws.cell(1, 1).value)

name_list = []
birth_date_list = []
number_list = []

for i in range(1, load_ws.max_row + 1):
    name_list.append(load_ws.cell(i, 1).value)
    birth_date_list.append(load_ws.cell(i, 2).value)
    number_list.append(load_ws.cell(i, 3).value)

print(name_list)
print(birth_date_list)
print(number_list)
